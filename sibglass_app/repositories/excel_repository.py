from __future__ import annotations

from pathlib import Path



class ExcelRepository:
    def read_lines(self, path: str) -> list[str]:
        return [" ".join(row).strip() for row in self.read_rows(path) if any(cell.strip() for cell in row)]

    def read_rows(self, path: str) -> list[list[str]]:
        suffix = Path(path).suffix.lower()
        if suffix == ".xlsx":
            return self._read_xlsx_rows(path)
        if suffix == ".xls":
            return self._read_xls_rows(path)
        raise ValueError("Поддерживаются только файлы .xlsx и .xls")

    def open_workbook(self, path: str):
        suffix = Path(path).suffix.lower()
        if suffix != ".xlsx":
            raise ValueError(
                "Файл заявки должен быть в формате .xlsx. Для записи в .xls сохраните шаблон как .xlsx и выберите его."
            )
        from openpyxl import load_workbook

        return load_workbook(path)

    @staticmethod
    def _read_xlsx_rows(path: str) -> list[list[str]]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=True)
        rows: list[list[str]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                rows.append([
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in row
                ])
        return rows

    @staticmethod
    def _read_xls_rows(path: str) -> list[list[str]]:
        try:
            import xlrd  # type: ignore
        except ImportError as exc:
            raise ValueError(
                "Для чтения файлов .xls установите зависимость xlrd>=2.0.1: pip install xlrd"
            ) from exc

        book = xlrd.open_workbook(path)
        rows: list[list[str]] = []
        for sheet in book.sheets():
            for row_idx in range(sheet.nrows):
                rows.append([
                    str(sheet.cell_value(row_idx, col_idx)).strip()
                    for col_idx in range(sheet.ncols)
                ])
        return rows
