from __future__ import annotations

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from sibglass_app.models.order_item import OrderItem
from sibglass_app.utils.excel_utils import find_cell_by_value


class SibglassWriterService:
    _BORDER_THIN = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    _FONT_DEFAULT = Font(name="Arial", size=10, color="000000")
    _FONT_ACCENT = Font(name="Arial", size=14, color="FF0000")
    _FILL_A = PatternFill(fill_type="solid", start_color="CCFFFF", end_color="CCFFFF")
    _ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    _ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

    def write(self, workbook, customer: str, address: str, items: list[OrderItem]) -> None:
        sheet = workbook.active
        self._fill_requisites(sheet, customer, address)
        self._write_items(sheet, items)

    @classmethod
    def _fill_requisites(cls, sheet: Worksheet, customer: str, address: str) -> None:
        customer_cell = find_cell_by_value(sheet, "Заказчик")
        address_cell = find_cell_by_value(sheet, "Адрес доставки")

        if customer_cell:
            cls._write_text_right_of_label(sheet, customer_cell.row, customer_cell.column, customer)

        if address_cell:
            cls._write_text_right_of_label(sheet, address_cell.row, address_cell.column, address)
            # Требование: строка адреса высотой 22px
            sheet.row_dimensions[address_cell.row].height = 22

    @classmethod
    def _write_text_right_of_label(cls, sheet: Worksheet, row: int, label_col: int, text: str) -> None:
        col = label_col + 1
        max_search = max(sheet.max_column + 20, col + 20)

        while col <= max_search:
            merged_range = cls._find_merged_range(sheet, row, col)
            if merged_range is None:
                cls._set_value_same_row_safe(sheet, row, col, text)
                return

            # Пишем только в merge-ячейку, якорь которой на этой же строке,
            # чтобы не ломать форматирование соседних строк
            anchor_row, anchor_col = merged_range.min_row, merged_range.min_col
            if anchor_row == row and anchor_col > label_col:
                cls._set_value_same_row_safe(sheet, anchor_row, anchor_col, text)
                return

            col = merged_range.max_col + 1

        cls._set_value_same_row_safe(sheet, row, label_col + 1, text)

    @staticmethod
    def _find_merged_range(sheet: Worksheet, row: int, col: int):
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return merged_range
        return None

    @classmethod
    def _set_value_same_row_safe(cls, sheet: Worksheet, row: int, col: int, value) -> None:
        cell_obj = sheet.cell(row=row, column=col)
        if isinstance(cell_obj, MergedCell):
            merged_range = cls._find_merged_range(sheet, row, col)
            if merged_range is None:
                return
            anchor_row, anchor_col = merged_range.min_row, merged_range.min_col
            if anchor_row != row:
                return
            sheet.cell(row=anchor_row, column=anchor_col, value=value)
            return

        sheet.cell(row=row, column=col, value=value)

    @classmethod
    def _set_value_safe(cls, sheet: Worksheet, row: int, col: int, value) -> None:
        cell_obj = sheet.cell(row=row, column=col)
        if isinstance(cell_obj, MergedCell):
            merged_range = cls._find_merged_range(sheet, row, col)
            if merged_range is None:
                return
            anchor_row, anchor_col = merged_range.min_row, merged_range.min_col
            if value in (None, "") and (anchor_row, anchor_col) != (row, col):
                return
            sheet.cell(row=anchor_row, column=anchor_col, value=value)
            return

        sheet.cell(row=row, column=col, value=value)

    @classmethod
    def _write_items(cls, sheet: Worksheet, items: list[OrderItem]) -> None:
        start_row, total_row = cls._find_table_bounds(sheet)
        existing_count = max(total_row - start_row, 0)
        target_count = len(items)

        if target_count > existing_count:
            sheet.insert_rows(total_row, target_count - existing_count)
            total_row += target_count - existing_count
        elif target_count < existing_count:
            sheet.delete_rows(start_row + target_count, existing_count - target_count)
            total_row -= existing_count - target_count

        for idx, item in enumerate(items, start=1):
            row = start_row + idx - 1
            cls._set_value_safe(sheet, row, 1, idx)
            cls._set_value_safe(sheet, row, 2, "")
            cls._set_value_safe(sheet, row, 3, item.formula)
            cls._set_value_safe(sheet, row, 4, int(item.width))
            cls._set_value_safe(sheet, row, 5, int(item.height))
            cls._set_value_safe(sheet, row, 6, int(item.count))
            cls._set_value_safe(sheet, row, 7, f"=D{row}*E{row}/1000000")
            cls._set_value_safe(sheet, row, 8, f"=G{row}*F{row}")
            cls._style_data_row(sheet, row)

        # Итого формулами
        if target_count > 0:
            cls._set_value_safe(sheet, total_row, 6, f"=SUM(F{start_row}:F{start_row + target_count - 1})")
            cls._set_value_safe(sheet, total_row, 7, f"=SUM(G{start_row}:G{start_row + target_count - 1})")
            cls._set_value_safe(sheet, total_row, 8, f"=SUM(H{start_row}:H{start_row + target_count - 1})")
        else:
            cls._set_value_safe(sheet, total_row, 6, 0)
            cls._set_value_safe(sheet, total_row, 7, 0)
            cls._set_value_safe(sheet, total_row, 8, 0)

        cls._style_total_cells(sheet, total_row)

    @classmethod
    def _style_data_row(cls, sheet: Worksheet, row: int) -> None:
        for col in range(1, 9):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                continue

            if col == 1:
                cell.fill = cls._FILL_A
                cell.border = cls._BORDER_THIN
                cell.font = cls._FONT_DEFAULT
            elif 2 <= col <= 6:
                cell.border = cls._BORDER_THIN
                cell.font = cls._FONT_DEFAULT
                if 3 <= col <= 6:
                    cell.alignment = cls._ALIGN_CENTER
                if col in (4, 5, 6):
                    cell.number_format = "0"
            else:  # G, H
                cell.border = cls._BORDER_THIN
                cell.font = cls._FONT_ACCENT
                cell.alignment = cls._ALIGN_RIGHT
                cell.number_format = "0.00"

    @classmethod
    def _style_total_cells(cls, sheet: Worksheet, row: int) -> None:
        for col in (6, 7, 8):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell, MergedCell):
                merged = cls._find_merged_range(sheet, row, col)
                if merged:
                    cell = sheet.cell(row=merged.min_row, column=merged.min_col)
            cell.font = cls._FONT_ACCENT
            if col == 6:
                cell.alignment = cls._ALIGN_CENTER
                cell.number_format = "0"
            else:
                cell.alignment = cls._ALIGN_RIGHT
                cell.number_format = "0.00"

    @staticmethod
    def _find_header_row(sheet: Worksheet) -> int | None:
        for row_idx in range(1, sheet.max_row + 1):
            for col in range(1, min(sheet.max_column, 8) + 1):
                value = sheet.cell(row_idx, col).value
                if str(value).strip() == "№":
                    return row_idx
        return None

    @staticmethod
    def _is_total_row(sheet: Worksheet, row_idx: int) -> bool:
        for col in (1, 2, 3):
            value = str(sheet.cell(row_idx, col).value or "").strip().lower()
            if "всего" in value:
                return True
        return False

    @classmethod
    def _find_table_bounds(cls, sheet: Worksheet) -> tuple[int, int]:
        header_row = cls._find_header_row(sheet)
        start_row = header_row + 1 if header_row is not None else 14

        total_row = None
        for row_idx in range(start_row, sheet.max_row + 1):
            if cls._is_total_row(sheet, row_idx):
                total_row = row_idx
                break

        if total_row is None:
            total_row = max(start_row, sheet.max_row + 1)

        return start_row, total_row
