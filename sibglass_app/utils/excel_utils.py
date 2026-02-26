from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet


def find_cell_by_value(sheet: Worksheet, needle: str):
    needle_lower = needle.strip().lower()
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            value = str(cell.value).strip().lower() if cell.value is not None else ""
            if value == needle_lower:
                return cell
    return None
